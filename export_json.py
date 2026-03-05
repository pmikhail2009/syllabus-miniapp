#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Экспорт данных из syllabuses.xlsx в frontend/data.json
Запуск: python export_json.py
"""
import openpyxl
import json
import os

EXCEL_FILE = 'syllabuses.xlsx'
OUTPUT_FILE = 'frontend/data.json'

def export_excel_to_json():
    if not os.path.exists(EXCEL_FILE):
        print(f"❌ Файл '{EXCEL_FILE}' не найден!")
        return False
    
    # Создаём папку frontend если нет
    os.makedirs('frontend', exist_ok=True)
    
    wb = openpyxl.load_workbook(EXCEL_FILE)
    data = {}
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        grade = sheet_name.replace('Класс ', '')
        data[grade] = {}
        
        for row in ws.iter_rows(min_row=2, values_only=False):
            if len(row) < 3:
                continue
            subject = row[0].value
            teacher = row[1].value
            url = row[2].value
            
            if not all([subject, teacher, url]):
                continue
            
            if subject not in data[grade]:
                data[grade][subject] = []
            
            data[grade][subject].append({
                'teacher': str(teacher).strip(),
                'url': str(url).strip()
            })
    
    wb.close()
    
    with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    
    print(f"✅ Экспорт завершён: {OUTPUT_FILE}")
    print(f"📊 Классов: {len(data)}, Предметов: {sum(len(s) for s in data.values())}")
    return True

if __name__ == '__main__':
    export_excel_to_json()