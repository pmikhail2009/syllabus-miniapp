# syllabus bot 
# сделал быстро, потом допиливаю, не судите строго ¯\_(ツ)_/¯

import telebot
from telebot import types
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os
from dotenv import load_dotenv

# грузим токен из .env 
load_dotenv()
TOKEN = os.getenv('BOT_TOKEN')
if not TOKEN:
    print("⚠️ нет токена в .env, бот не запустится")
    exit(1)

tg_bot = telebot.TeleBot(TOKEN)
XLS_PATH = 'syllabuses.xlsx'  

# храним, какой класс выбрал юзер (по chat_id)
user_grade = {}


def make_excel():
    """создаёт эксель с силлабусами если нет файла
    данные жёстко зашиты — лень выносить в конфиг, потом поправлю
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # удаляем пустой лист

    # стили для шапки — сделал один раз, потом копипаст
    h_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    h_font = Font(bold=True, color="FFFFFF", size=12)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # данные по классам — если надо добавить предмет, просто допишите в список
    # формат: (предмет, учитель, ссылка)
    data = {
        '8': [
            ('Русский язык', 'Е. О. Борисова', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_8_1_rus_b_beo_2025.pdf'),
            ('Литература', 'Ю. Ю. Ишутин', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_8_1_lit_b_iyy_2025.pdf'),
            ('История', 'И. Ю. Романов', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_8_1_ist_b_riy_2025.pdf'),
            ('Обществознание', 'А. Н. Егорова', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_8_1_obsh_b_ean_2025.pdf'),
            ('Право', 'А. В. Немова', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_8_1_pravo_b_nav_2025.pdf'),
            ('Английский язык', 'Д. В. Плотникова', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_8_1_eng_u_pdv_2025.pdf'),
            ('Английский язык', 'Е. В. Петрова', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_8_1_eng_u_pev_2025.pdf'),
            ('Английский язык', 'У. Райс', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_8_1_eng_u_ru_2025.pdf'),
            ('Испанский язык', 'М. М. Шустовских', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_8_1_isp_b_1_shmm_2025.pdf'),
            ('Китайский язык', 'Н. Н. Шаталова', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_8_1_kit_b_shnn_2025.pdf'),
            ('География', 'Е. В. Бетехтин', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_8_1_geo_b_bev_2025.pdf'),
            ('Экономика', 'И. А. Нимерницкая', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_8_1_econ_b_nia_2025.pdf'),
            ('Биология', 'В. В. Майчак', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_8_1_bio_b_mvv_2025.pdf'),
            ('Физика', 'А. А. Кочегаров', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_8_1_fiz_b_kaa_2025.pdf'),
            ('Химия', 'Д. В. Ефанов', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_8_1_him_b_edv_2025.pdf'),
            ('Алгебра', 'А. А. Коваленко', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_8_1_alg_b_kaa_2025.pdf'),
            ('Геометрия', 'А. А. Коваленко', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_8_1_geom_b_kaa_2025.pdf'),
            ('Теория вероятности', 'А. А. Коваленко', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_8_1_teorver_b_kaa_2025.pdf'),
            ('Информатика', 'Д.А. Панфилова, А.Ю. Величко', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_8_1_inf_b_pda_vay_2025.pdf'),
            ('Технология', 'Д. А. Панфилова', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_8_1_tekn_b_pda_2025.pdf'),
        ],
        '9': [
            ('Русский язык', 'Е. О. Борисова', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_9_1_rus_b_beo_2025.pdf'),
            ('Литература', 'Ю. Ю. Ишутин', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_9_1_lit_b_iyy_2025.pdf'),
            ('История', 'И. Ю. Романов', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_9_1_ist_b_riy_2025.pdf'),
            ('Обществознание', 'А. Н. Егорова', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_9_1_obsh_b_ean_2025.pdf'),
            ('Право', 'А. В. Немова', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_9_1_pravo_b_nav_2025.pdf'),
            ('Теория познания', 'С. М. Гнездилов', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_9_1_tok_b_gsm_2025.pdf'),
            ('Английский язык', 'У. Райс (1)', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_9_1_eng_u_ru_2025.pdf'),
            ('Английский язык', 'У. Райс (2)', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_9_1_eng_u_2_ru_2025.pdf'),
            ('Английский язык', 'Д. В. Плотникова', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_9_1_eng_u_pdv_2025.pdf'),
            ('Английский язык', 'В. В. Хитрук (GWB1)', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_9_1_eng_u_GWB1_hvv_2025.pdf'),
            ('Английский язык', 'В. В. Хитрук (GWB1+)', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_9_1_eng_u_GWB1+_hvv_2025.pdf'),
            ('Английский язык', 'Е. Петрова', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_9_1_eng_u_pev_2025.pdf'),
            ('Немецкий язык', 'О. Б. Темежникова (начинающие)', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_9_1_nem_n_tob_2025.pdf'),
            ('Немецкий язык', 'О. Б. Темежникова (продолжающие-1)', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_9_1_nem_p1_tob_2025.pdf'),
            ('Немецкий язык', 'О. Б. Темежникова (продолжающие-2)', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_9_1_nem_p2_tob_2025.pdf'),
            ('Испанский язык', 'М. М. Шустовских (1)', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_9_1_isp_b_1_shmm_2025.pdf'),
            ('Испанский язык', 'М. М. Шустовских (2)', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_9_1_isp_b_2_shmm_2025.pdf'),
            ('География', 'Е. В. Бетехтин', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_9_1_geo_b_bev_2025.pdf'),
            ('Экономика', 'И. А. Нимерницкая', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_9_1_econ_b_nia_2025.pdf'),
            ('Биология', 'В. В. Майчак', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_9_1_bio_b_mvv_2025.pdf'),
            ('Физика', 'В. Н. Белянин', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_9_1_fiz_b_bvn_2025.pdf'),
            ('Химия', 'Д. В. Ефанов', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_9_1_him_b_edv_2025.pdf'),
            ('Алгебра', 'А. В. Пчелинцева', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_9_1_alg_b_pav_2025.pdf'),
            ('Геометрия', 'А. В. Пчелинцева', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_9_1_geom_b_pav_2025.pdf'),
            ('Теория вероятности', 'А. В. Пчелинцева', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_9_1_teorver_b_pav_2025.pdf'),
            ('Информатика', 'Д.А. Панфилова, А.Ю. Величко', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_9_1_inf_b_pda_vay_2025.pdf'),
            ('Технология', 'Д. А. Панфилова', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_9_1_tekn_b_pda_2025.pdf'),
        ],
        '10': [
            ('Русский язык', 'Е. А. Пантелеева', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_10_1_rus_b_pea_2025.pdf'),
            ('Литература', 'П. Ф. Подковыркин (базовый)', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_10_1_lit_b_ppf_2025.pdf'),
            ('Литература', 'П. Ф. Подковыркин (углублённый)', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_10_1_lit_u_ppf_2025.pdf'),
            ('История', 'А. Н. Кадыков', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_10_1_ist_b_kan_2025.pdf'),
            ('История', 'В. К. Герасимов', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_10_1_ist_u_gvk_2025.pdf'),
            ('Обществознание', 'С. М. Гнездилов', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_10_1_obsh_b_gsm_2025.pdf'),
            ('Обществознание', 'А. Н. Егорова', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_10_1_obsh_b_ean_2025.pdf'),
            ('Право', 'А. В. Юсупов', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_10_1_prav_u_uav_2025.pdf'),
            ('Право', 'С.Н. Мореева', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_10_1_prav_u_msn_2025.pdf'),
            ('Теория и история искусства', 'Д. С. Матюнина', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_10_1_teor_isk_b_vkn_2025.pdf'),
            ('Теория познания', 'И. С. Курилович', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_10_1_ToK_u_kis_2025.pdf'),
            ('Английский язык', 'С. С. Московская', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_10_1_eng_u_mss_2025.pdf'),
            ('Английский язык', 'Е. В. Петрова', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_10_1_eng_u_pev_2025.pdf'),
            ('Английский язык', 'Д. В. Плотникова', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_10_1_eng_u_pdv_2025.pdf'),
            ('Немецкий язык', 'О. Б. Темежникова (начинающие)', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_10_1_nem_n_tob_2025.pdf'),
            ('Немецкий язык', 'О. Б. Темежникова (продолжающие-1)', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_10_1_nem_p1_tob_2025.pdf'),
            ('Немецкий язык', 'О. Б. Темежникова (продолжающие-2)', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_10_1_nem_p2_tob_2025.pdf'),
            ('Немецкий язык', 'О. Б. Темежникова (продолжающие-3)', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_10_1_nem_p3_tob_2025.pdf'),
            ('Испанский язык', 'М. М. Шустовских (1)', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_10_1_isp_b_1_shmm_2025.pdf'),
            ('Испанский язык', 'М. М. Шустовских (2)', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_10_1_isp_b_2_shmm_2025.pdf'),
            ('Испанский язык', 'М. М. Шустовских (3)', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_10_1_isp_b_3_shmm_2025.pdf'),
            ('Испанский язык', 'М. М. Шустовских (4)', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_10_1_isp_b_4_shmm_2025.pdf'),
            ('Итальянский язык', 'Е. С. Гурьева', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_10_1_ital_b_ges_2025.pdf'),
            ('Китайский язык', 'Н. Н. Шаталова', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_10_1_kit_b_shnn_2025.pdf'),
            ('Арабский язык', 'Е. И. Смирнова', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_10_1_arab_b_sei_2025.pdf'),
            ('География', 'Д. В. Степаньков', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_10_1_geo_b_sdv_2025.pdf'),
            ('Экономика', 'И. А. Нимерницкая (базовый)', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_10_1_econ_b_nia_2025.pdf'),
            ('Экономика', 'И. А. Нимерницкая (углублённый)', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_10_1_econ_u_nia_2025.pdf'),
            ('Экономика', 'М. И. Мозгачёв', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_10_1_econ_u_mmi_2025.pdf'),
            ('Биология', 'В. В. Майчак (базовый)', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_10_1_bio_b_mvv_2025.pdf'),
            ('Биология', 'В. В. Майчак (углублённый)', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_10_1_bio_u_mvv_2025.pdf'),
            ('Химия', 'Д. В. Ефанов', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_10_1_him_b_edv_2025.pdf'),
            ('Физика', 'А. А. Кочегаров', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_10_1_fiz_b_kaa_2025.pdf'),
            ('Алгебра', 'А. М. Городнов (базовый)', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_10_1_alg_b_gam_2025.pdf'),
            ('Алгебра', 'В. К. Ушаков (углублённый)', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_10_1_alg_u_uvk_2025.pdf'),
            ('Алгебра', 'А. Ю. Величко (углублённый)', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_10_1_alg_u_vay_2025.pdf'),
            ('Алгебра', 'А. А. Коваленко (базовый)', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_10_1_alg_b_kaa_2025.pdf'),
            ('Геометрия', 'А. М. Городнов (базовый)', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_10_1_geom_b_gam_2025.pdf'),
            ('Геометрия', 'В. К. Ушаков (углублённый)', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_10_1_geom_u_uvk_2025.pdf'),
            ('Геометрия', 'А. Ю. Величко (углублённый)', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_10_1_geom_u_vay_2025.pdf'),
            ('Геометрия', 'А. А. Коваленко (базовый)', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_10_1_geom_b_kaa_2025.pdf'),
            ('Информатика', 'Д. А. Панфилова, А. Ю. Величко', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_10_1_inf_u_pda_vay_2025.pdf'),
        ],
        '11': [
            ('Русский язык', 'Е. В. Гаранина, Е. О. Борисова', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_11_1_rus_b_gev_beo_2025.pdf'),
            ('Литература', 'А. С. Сенилова (базовый)', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_11_1_lit_b_sas_2025.pdf'),
            ('Литература', 'Е. В. Гаранина, Ю. Ю. Ишутин (углублённый)', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_11_1_lit_u_gev_iyy_2025.pdf'),
            ('Теория и история искусства', 'Д. С. Матюнина', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_11_1_teor_isk_b_mds_2025.pdf'),
            ('История', 'А. Н. Кадыков (базовый)', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_11_1_ist_b_kan_2025.pdf'),
            ('История', 'А. Ф. Цветкова (углублённый)', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_11_1_ist_u_caf_2025.pdf'),
            ('Всеобщая политическая история', 'А. А. Атаманенко', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_11_1_polist_u_aaa_2025.pdf'),
            ('Политология', 'С. А. Кожеуров', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_11_1_polit_u_ksa_2025.pdf'),
            ('Право', 'А. В. Юсупов', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_11_1_prav_u_uav_2025.pdf'),
            ('Право', 'С.Н. Мореева (юридический)', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_11_1_prav_u_2_msn_2025.pdf'),
            ('Обществознание', 'А. Н. Егорова (базовый)', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_11_1_obsh_b_ean_2025.pdf'),
            ('Обществознание', 'А. Н. Егорова (углублённый)', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_11_1_obsh_u_ean_2025.pdf'),
            ('Социология', 'М. А. Серкина', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_11_1_soc_u_sma_2025.pdf'),
            ('Английский язык', 'В. В. Хитрук (Gateway B2)', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_11_1_eng_u_GWB2_hvv_2025.pdf'),
            ('Английский язык', 'В. В. Хитрук (Gateway C1)', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_11_1_eng_u_GWC1_hvv_2025.pdf'),
            ('Английский язык', 'В. В. Хитрук (Prepare 7)', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_11_1_eng_u_PR7_hvv_2025.pdf'),
            ('Английский язык', 'С. С. Московская', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_11_1_eng_u_mss_2025.pdf'),
            ('Английский язык', 'А. Л. Ромашкина', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_11_1_eng_u_ral_2025.pdf'),
            ('Английский язык', 'Е. С. Гурина', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_11_1_eng_u_1_ges_2025.pdf'),
            ('Английский язык', 'У. Райс', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_11_1_eng_u_ru_2025.pdf'),
            ('Английский язык State Exam', 'Е. В. Петрова', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_11_1_state_u_pev_2025.pdf'),
            ('Английский язык State Exam', 'С. С. Московская', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_11_1_state_u_mss_2025.pdf'),
            ('Немецкий язык', 'О. Б. Темежникова (продолжающие-1)', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_11_1_nem_p1_tob_2025.pdf'),
            ('Немецкий язык', 'О. Б. Темежникова (продолжающие-2)', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_11_1_nem_p2_tob_2025.pdf'),
            ('Немецкий язык', 'О. Б. Темежникова (продолжающие-3)', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_11_1_nem_p3_tob_2025.pdf'),
            ('Испанский язык', 'Е. В. Львова', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_11_1_isp_b_lev_2025.pdf'),
            ('Испанский язык', 'М. М. Шустовских (3)', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_11_1_isp_b_3_shmm_2025.pdf'),
            ('Испанский язык', 'М. М. Шустовских (4)', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_11_1_isp_b_4_shmm_2025.pdf'),
            ('Итальянский язык', 'Е. С. Гурьева', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_11_1_ital_b_ges_2025.pdf'),
            ('Французский язык', 'М. С. Салкина', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_11_1_fr_b_sms_2025.pdf'),
            ('Китайский язык', 'Н. Н. Шаталова', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_11_1_kit_u_shnn_2025.pdf'),
            ('Арабский язык', 'Е. И. Смирнова', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_11_1_arab_b_sei_2025.pdf'),
            ('География', 'Д. В. Степаньков', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_11_1_geo_b_sdv_2025.pdf'),
            ('Экономика', 'И. А. Нимерницкая (базовый)', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_11_1_econ_b_nia_2025.pdf'),
            ('Экономика', 'И. А. Нимерницкая (углублённый)', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_11_1_econ_u_nia_2025.pdf'),
            ('Биология', 'В. В. Майчак', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_11_1_bio_u_mvv_2025.pdf'),
            ('Алгебра', 'В. К. Ушаков', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_11_1_alg_u_uvk_2025.pdf'),
            ('Алгебра', 'В. С. Осипова', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_11_1_alg_u_ovs_2025.pdf'),
            ('Алгебра', 'А. В. Петров', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_11_1_alg_u_pav_2025.pdf'),
            ('Геометрия', 'В. К. Ушаков', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_11_1_geom_u_uvk_2025.pdf'),
            ('Геометрия', 'В. С. Осипова', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_11_1_geom_u_ovs_2025.pdf'),
            ('Геометрия', 'А. В. Петров', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_11_1_geom_u_pav_2025.pdf'),
            ('Информатика', 'Ю. П. Мартемьянов (базовый)', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_11_1_inf_b_myp_2025.pdf'),
            ('Информатика', 'Д. А. Панфилова, А. Ю. Величко (углублённый)', 'https://ranepa-lyceum.ru/docs/sil_25-26_1/sil_11_1_inf_u_pda_vay_2025.pdf'),
        ],
    }

    for cls, rows in data.items():
        sheet = wb.create_sheet(f'Класс {cls}')
        sheet['A1'], sheet['B1'], sheet['C1'] = 'Предмет', 'Учитель', 'Ссылка на силлабус'
        for c in ['A1', 'B1', 'C1']:
            sheet[c].fill, sheet[c].font, sheet[c].border = h_fill, h_font, thin_border
        sheet.column_dimensions['A'].width, sheet.column_dimensions['B'].width, sheet.column_dimensions['C'].width = 30, 25, 50

        r = 2
        for subj, teach, link in rows:
            sheet[f'A{r}'], sheet[f'B{r}'], sheet[f'C{r}'] = subj, teach, link
            for cell in [f'A{r}', f'B{r}', f'C{r}']:
                sheet[cell].border = thin_border
                sheet[cell].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            r += 1
        sheet.row_dimensions[1].height = 25
        for i in range(2, r):
            sheet.row_dimensions[i].height = 30

    wb.save(XLS_PATH)
    print(f"✓ эксель создан: {XLS_PATH}")


def load_data():
    """читает силлабусы из экселя
    если файла нет — создаёт новый с дефолтными данными
    возвращает словарь: {класс: {предмет: [{учитель, ссылка}]}}
    """
    if not os.path.exists(XLS_PATH):
        print("файла нет, создаю...")
        make_excel()

    result = {}
    wb = openpyxl.load_workbook(XLS_PATH)
    for sh in wb.sheetnames:
        cls = sh.replace('Класс ', '')
        result[cls] = {}
        ws = wb[sh]
        for row in ws.iter_rows(min_row=2, values_only=False):
            if len(row) < 3 or not all([row[0].value, row[1].value, row[2].value]):
                continue
            subj, teach, link = row[0].value, row[1].value, row[2].value
            if subj not in result[cls]:
                result[cls][subj] = []
            result[cls][subj].append({'teacher': teach, 'url': link})
    wb.close()
    return result


# клавиатуры — вынес отдельно, чтобы не дублировать код
def main_kb():
    mk = types.ReplyKeyboardMarkup(resize_keyboard=True)
    mk.add('8 класс', '9 класс')
    mk.add('10 класс', '11 класс')
    mk.add('❓ Справка', '🎓 Mini App')
    return mk


def subjects_kb(grade, db):
    mk = types.ReplyKeyboardMarkup(resize_keyboard=True)
    if grade in db:
        subs = sorted(db[grade].keys())
        for i in range(0, len(subs), 2):
            mk.add(subs[i], subs[i+1]) if i+1 < len(subs) else mk.add(subs[i])
    mk.add('⬅️ Назад')
    return mk


def teachers_inline(teachers):
    mk = types.InlineKeyboardMarkup()
    for i, t in enumerate(teachers):
        mk.add(types.InlineKeyboardButton(t['teacher'], callback_data=f"t_{i}"))
    return mk


# загружаем данные при старте
db = load_data()


# === обработчики команд ===
@tg_bot.message_handler(commands=['start'])
def start(msg):
    user_grade[msg.chat.id] = None
    tg_bot.send_message(
        msg.chat.id,
        "👋 Привет! Я бот с силлабусами лицея РАНХиГС.\nВыбери класс ниже 👇",
        reply_markup=main_kb()
    )


@tg_bot.message_handler(commands=['help'])
def help_msg(msg):
    tg_bot.send_message(
        msg.chat.id,
        "📋 Как пользоваться:\n"
        "1. Нажми на класс (8–11)\n"
        "2. Выбери предмет\n"
        "3. Если учителей несколько — выбери нужного\n"
        "4. Получи ссылку на силлабус\n\n"
        "Команды:\n"
        "/start — меню\n"
        "/help — эта справка\n"
        "/reload — обновить данные из Excel\n"
        "/webapp — открыть мини-приложение\n"
        "/rep — перейти в репозиторий на GitHub",
        reply_markup=main_kb()
    )


@tg_bot.message_handler(commands=['reload'])
def reload_db(msg):
    global db
    try:
        db = load_data()
        tg_bot.send_message(msg.chat.id, "✓ данные обновлены", reply_markup=main_kb())
    except Exception as e:
        tg_bot.send_message(msg.chat.id, f"✗ ошибка: {e}", reply_markup=main_kb())


@tg_bot.message_handler(commands=['webapp'])
def send_webapp(msg):
    url = "https://pmikhail2009.github.io/syllabus-miniapp/"
    mk = types.InlineKeyboardMarkup()
    mk.add(types.InlineKeyboardButton("🎓 Открыть силлабусы", web_app=types.WebAppInfo(url=url)))
    tg_bot.send_message(msg.chat.id, "📱 Мини-приложение:", reply_markup=mk)


@tg_bot.message_handler(commands=['rep'])
def send_repo(msg):
    url = "https://github.com/pmikhail2009/syllabus-miniapp"
    mk = types.InlineKeyboardMarkup()
    mk.add(types.InlineKeyboardButton("🔗 GitHub", url=url))
    tg_bot.send_message(msg.chat.id, f"📦 Исходный код:\n{url}", reply_markup=mk)


# === обработчики кнопок меню ===
@tg_bot.message_handler(func=lambda m: m.text in ['8 класс', '9 класс', '10 класс', '11 класс'])
def pick_grade(msg):
    cid = msg.chat.id
    grade = msg.text.split()[0]  
    user_grade[cid] = grade

    if grade not in db:
        tg_bot.send_message(cid, "✗ нет данных для этого класса", reply_markup=main_kb())
        return

    count = len(db[grade])
    tg_bot.send_message(
        cid,
        f"📚 {grade} класс ({count} предметов)\nВыбери предмет:",
        reply_markup=subjects_kb(grade, db)
    )


@tg_bot.message_handler(func=lambda m: m.text == '❓ Справка')
def btn_help(msg):
    help_msg(msg)


@tg_bot.message_handler(func=lambda m: m.text == '🎓 Mini App')
def btn_webapp(msg):
    send_webapp(msg)


@tg_bot.message_handler(func=lambda m: m.text == '⬅️ Назад')
def go_back(msg):
    user_grade[msg.chat.id] = None
    tg_bot.send_message(msg.chat.id, "📌 Выбери класс:", reply_markup=main_kb())


# === выбор предмета ===
@tg_bot.message_handler(func=lambda m: True)
def pick_subject(msg):
    cid = msg.chat.id
    subj = msg.text
    grade = user_grade.get(cid)

    if not grade or grade not in db or subj not in db[grade]:
        tg_bot.send_message(cid, "✗ используй кнопки меню", reply_markup=main_kb())
        return

    teachers = db[grade][subj]

    # если учитель один — сразу отдаём ссылку
    if len(teachers) == 1:
        t = teachers[0]
        tg_bot.send_message(
            cid,
            f"✅ {subj}\n📍 {grade} класс\n👨‍🏫 {t['teacher']}\n\n🔗 {t['url']}",
            reply_markup=main_kb()
        )
        user_grade[cid] = None
        return

    # если несколько — показываем выбор
    user_grade[cid] = {'grade': grade, 'subj': subj, 'teachers': teachers}
    tg_bot.send_message(
        cid,
        f"👨‍🏫 Выбери учителя по предмету '{subj}':",
        reply_markup=teachers_inline(teachers)
    )


# === callback: выбор учителя ===
@tg_bot.callback_query_handler(func=lambda c: c.data.startswith('t_'))
def on_teacher_pick(call):
    cid = call.message.chat.id
    idx = int(call.data.split('_')[1])
    ctx = user_grade.get(cid)

    if not ctx or not isinstance(ctx, dict):
        tg_bot.answer_callback_query(call.id, "сессия истекла, начни заново")
        return

    grade, subj, teachers = ctx['grade'], ctx['subj'], ctx['teachers']
    if idx >= len(teachers):
        tg_bot.answer_callback_query(call.id, "ошибка выбора")
        return

    t = teachers[idx]
    tg_bot.delete_message(cid, call.message.message_id)
    tg_bot.send_message(
        cid,
        f"✅ {subj}\n📍 {grade} класс\n👨‍🏫 {t['teacher']}\n\n🔗 {t['url']}",
        reply_markup=main_kb()
    )
    user_grade[cid] = None
    tg_bot.answer_callback_query(call.id)


# === запуск ===
if __name__ == '__main__':
    print(f"🤖 бот запущен | файл: {XLS_PATH} | классов: {len(db)}")
    print("команды: /start /help /reload /webapp /rep")
    try:
        tg_bot.infinity_polling()
    except KeyboardInterrupt:
        print("\n⏹ остановлен")
    except Exception as e:
        print(f"✗ упал: {e}")