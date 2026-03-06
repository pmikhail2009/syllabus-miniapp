# scraper/excel_builder.py
# Создаёт или обновляет Excel-файл с силлабусами

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
from config import EXCEL_PATH


def create_or_update_excel(syllabus_list):
    """
    Создаёт новый Excel или обновляет существующий.
    Структура листа: Предмет | Учитель | Ссылка
    """
    # если файла нет — создаём пустой шаблон
    if not os.path.exists(EXCEL_PATH):
        _create_template()
    
    wb = openpyxl.load_workbook(EXCEL_PATH)
    
    # группируем по классам
    by_grade = {}
    for item in syllabus_list:
        g = item['grade']
        if g not in by_grade:
            by_grade[g] = []
        by_grade[g].append(item)
    
    for grade, items in by_grade.items():
        sheet_name = f'Класс {grade}'
        
        # создаём лист если нет
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(sheet_name)
            _setup_header(ws)
        else:
            ws = wb[sheet_name]
            # чистим старые данные (шапку оставляем)
            _clear_sheet(ws)
        
        # заполняем
        _fill_sheet(ws, items)
    
    wb.save(EXCEL_PATH)
    print(f"✓ Excel обновлён: {EXCEL_PATH}")


def _create_template():
    """Создаёт пустой шаблон с листами для 8-11 классов"""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for g in ['8', '9', '10', '11']:
        ws = wb.create_sheet(f'Класс {g}')
        _setup_header(ws)
    wb.save(EXCEL_PATH)


def _setup_header(ws):
    """Настраивает шапку листа"""
    ws['A1'], ws['B1'], ws['C1'] = 'Предмет', 'Учитель', 'Ссылка на силлабус'
    
    h_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    h_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    for c in ['A1', 'B1', 'C1']:
        ws[c].fill = h_fill
        ws[c].font = h_font
        ws[c].border = border
    
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 60


def _clear_sheet(ws):
    """Удаляет все строки кроме шапки"""
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=3):
        for cell in row:
            cell.value = None


def _fill_sheet(ws, items):
    """Заполняет лист данными"""
    row = 2
    for item in items:
        ws[f'A{row}'] = item['subject']
        ws[f'B{row}'] = item['teacher']
        ws[f'C{row}'] = item['url']
        
        # стиль ячеек
        for col in ['A', 'B', 'C']:
            cell = ws[f'{col}{row}']
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
        row += 1
    
    # высота строк
    ws.row_dimensions[1].height = 25
    for i in range(2, row):
        ws.row_dimensions[i].height = 30